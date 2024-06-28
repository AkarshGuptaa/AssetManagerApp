import tkinter as tk
from tkinter import ttk
import openpyxl
from tkcalendar import DateEntry, Calendar

root = tk.Tk()
root.title('Asset Manager')

style = ttk.Style(root)
root.tk.call("source", "forest-dark.tcl")
root.tk.call("source", "forest-light.tcl")
style.theme_use("forest-light")


def toggle_mode():
    if mode_switch.instate(['selected']):
        style.theme_use('forest-dark')
    else:
        style.theme_use('forest-light')


def load_data():
    global list_values
    path = "file.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)

    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)
    reload_treeview(list_values[1:])


def insert_row():
    asset = asset_entry.get()
    model = model_combobox.get()
    user = user_entry.get()
    sn = sn_entry.get()
    monitor = monitor_entry.get()
    eport = eport_entry.get()
    keyboard = keyboard_entry.get()
    cal = cal_entry.get()

    # Insert row into Excel sheet
    path = 'file.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [asset, model, user, sn, monitor, eport, keyboard, cal]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert('', tk.END, values=row_values)

    # Clear the values
    asset_entry.delete(0, 'end')
    asset_entry.insert(0, 'Asset#')
    model_combobox.set(model_list[0])
    user_entry.delete(0, 'end')
    user_entry.insert(0, 'User')
    sn_entry.delete(0, 'end')
    sn_entry.insert(0, 'S/N')
    monitor_entry.delete(0, 'end')
    monitor_entry.insert(0, 'Monitor S/N')
    eport_entry.delete(0, 'end')
    eport_entry.insert(0, 'E Port')
    keyboard_entry.delete(0, 'end')
    keyboard_entry.insert(0, 'Keyboard')


def reload_treeview(data):
    treeview.delete(*treeview.get_children())
    for row_values in data:
        treeview.insert("", tk.END, values=row_values)


def clear_search():
    search_entry.delete(0, tk.END)
    reload_treeview(list_values[1:])


def search(event=None):
    search_term = search_entry.get().lower()
    selected_column = column_combobox.get()
    column_index = list_values[0].index(selected_column) if selected_column in list_values[0] else None

    if column_index is not None:
        filtered_data = [
            row for row in list_values[1:]
            if search_term in str(row[column_index]).lower()
        ]
        reload_treeview(filtered_data)


def edit_row(event):
    selected_item = treeview.selection()[0]
    selected_values = treeview.item(selected_item, "values")

    edit_popup = tk.Toplevel(root)
    edit_popup.title("Edit Row")

    entries = []
    labels = list_values[0]

    for i, value in enumerate(selected_values):
        ttk.Label(edit_popup, text=labels[i]).grid(row=i, column=0, padx=5, pady=5)
        entry = ttk.Entry(edit_popup)
        entry.insert(0, value)
        entry.grid(row=i, column=1, padx=5, pady=5)
        entries.append(entry)

    def save_changes():
        new_values = [entry.get() for entry in entries]
        treeview.item(selected_item, values=new_values)

        # Update the Excel file
        path = 'file.xlsx'
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        row_index = treeview.index(selected_item) + 2
        for col_index, new_value in enumerate(new_values, start=1):
            sheet.cell(row=row_index, column=col_index).value = new_value
        workbook.save(path)

        edit_popup.destroy()

    ttk.Button(edit_popup, text="Save", command=save_changes).grid(row=len(selected_values), column=0, columnspan=2,
                                                                   pady=10)


frame = ttk.Frame(root)
frame.pack()

widget_frame = ttk.LabelFrame(frame, text="Insert Row")
widget_frame.grid(row=0, column=0, padx=20, pady=10)

asset_entry = ttk.Entry(widget_frame)
asset_entry.insert(0, 'Asset#')
asset_entry.bind("<FocusIn>", lambda e: asset_entry.delete('0', 'end'))
asset_entry.grid(row=0, column=0, padx=5, pady=(0, 5), sticky='ew')

model_list = ['Model', 'Latitude E5420', 'Latitude E6230', 'Latitude E5440', 'Latitude 6410', 'Latitude E6540']

model_combobox = ttk.Combobox(widget_frame, values=model_list)
model_combobox.current(0)
model_combobox.grid(row=1, column=0, padx=5, pady=5, sticky='ew')

user_entry = ttk.Entry(widget_frame)
user_entry.insert(0, "User")
user_entry.bind("<FocusIn>", lambda e: user_entry.delete('0', 'end'))
user_entry.grid(row=2, column=0, padx=5, pady=5, sticky='ew')

sn_entry = ttk.Entry(widget_frame)
sn_entry.insert(0, "S/N:")
sn_entry.bind("<FocusIn>", lambda e: sn_entry.delete('0', 'end'))
sn_entry.grid(row=3, column=0, padx=5, pady=5, sticky='ew')

monitor_entry = ttk.Entry(widget_frame)
monitor_entry.insert(0, "Monitor SN")
monitor_entry.bind("<FocusIn>", lambda e: monitor_entry.delete('0', 'end'))
monitor_entry.grid(row=4, column=0, padx=5, pady=5, sticky='ew')

eport_entry = ttk.Entry(widget_frame)
eport_entry.insert(0, "E Port")
eport_entry.bind("<FocusIn>", lambda e: eport_entry.delete('0', 'end'))
eport_entry.grid(row=5, column=0, padx=5, pady=5, sticky='ew')

keyboard_entry = ttk.Entry(widget_frame)
keyboard_entry.insert(0, "Keyboard")
keyboard_entry.bind("<FocusIn>", lambda e: keyboard_entry.delete('0', 'end'))
keyboard_entry.grid(row=6, column=0, padx=5, pady=5, sticky='ew')

cal_label = ttk.LabelFrame(widget_frame, text='Purchase Date')
cal_label.grid(row=7, column=0, padx=5, pady=5, sticky='ew')
cal_entry = DateEntry(cal_label, date_pattern='dd/mm/yyyy')
cal_entry.grid(row=8, column=0, padx=5, pady=5, sticky='nsew')

button = ttk.Button(widget_frame, text="Insert", command=insert_row)
button.grid(row=9, column=0, padx=5, pady=5, sticky='nsew')

seperator = ttk.Separator(widget_frame)
seperator.grid(row=10, column=0, padx=5, pady=5, sticky='nsew')

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)

path = "file.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
list_values = list(sheet.values)

search_label = ttk.LabelFrame(widget_frame, text='Filter')
search_label.grid(row=11, column=0, padx=5, pady=5, sticky="nsew")

column_combobox = ttk.Combobox(search_label, values=list_values[0])
column_combobox.grid(row=12, column=0, padx=5, pady=5, sticky="nsew")
column_combobox.current(0)

search_entry = ttk.Entry(search_label)
search_entry.grid(row=13, column=0, padx=5, pady=5, sticky="nsew")
search_entry.bind("<KeyRelease>", search)

seperator2 = ttk.Separator(widget_frame)
seperator2.grid(row=14, column=0, padx=5, pady=5, sticky='ew')

mode_switch = ttk.Checkbutton(widget_frame, text='Mode', style='Switch', command=toggle_mode)
mode_switch.grid(row=15, column=0, padx=5, pady=5, sticky='nsew')

treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side='right', fill='y')

cols = ("Asset#", "Model No", "User", "S/N", "Monitor SN", "E Port", "Keyboard", "Purchase Date")

treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=40)
treeview.column("Asset#", width=130)
treeview.column("Model No", width=130)
treeview.column("User", width=100)
treeview.column("S/N", width=230)
treeview.column("Monitor SN", width=230)
treeview.column("E Port", width=130)
treeview.column("Keyboard", width=130)
treeview.column("Purchase Date", width=100)
treeview.pack()

treeview.bind("<Double-1>", edit_row)
treeScroll.config(command=treeview.yview)
load_data()

root.mainloop()
